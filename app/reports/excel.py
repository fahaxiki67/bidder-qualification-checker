"""Excel 核查明细表（11 sheet，任务书 §16 表结构一一对应 + 汇总封面）。

数据口径：最新一次【完整】核查运行（check_runs.finished_at 非空），与 Web 结果页一致；
历史批次不混入。全部状态经 report_label 输出——"查询失败"绝不写成"无异常"。
"""
from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill

from .. import __version__
from ..core.status import Status, report_label

HEADER_FONT = Font(bold=True)

#: 公式注入风险前缀（= + - @ 及制表/回车）：报告内容大量来自第三方页面文本
#: （企业名称、处罚描述、名单备注…），落单元格前一律加前导单引号强制按文本处理，
#: 绝不给表格软件/CSV 再导出留下"把文本当公式执行"的机会。
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _source_provenance(run) -> str:
    """数据来源口径：mock 演示链路绝不能被当成真实官方查询结果使用。"""
    if not run:
        return "尚无完整核查运行"
    if run["scenario"] == "real_sources":
        return "真实官方数据源查询（real_sources）"
    return (f"mock 演示链路（场景 {run['scenario']}）：非真实官方查询，"
            "仅用于流程演示，不得作为核查结论使用")


def _cell(value):
    """单元格消毒：危险前缀前置单引号（Excel/WPS/LibreOffice 均按文本显示，不显示引号本身）。"""
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value


#: 报告内置公式的唯一落点：封面与汇总的统计区块。
#: 公式一律由程序按行号构造，绝不拼接企业名/处罚描述等第三方文本，
#: 与 _cell() 消毒形成两条互斥路径——数据路径零公式，公式路径零外部文本。
_STAT_SHEET = "封面与汇总"

#: 状态列高亮（条件格式的公式规则）：红=否决/失败，黄=需人工/风险，绿=通过。
_STATUS_HIGHLIGHTS = (
    ("FAIL", "FFC7CE", "9C0006"),
    ("ERROR", "FFC7CE", "9C0006"),
    ("TIMEOUT", "FFC7CE", "9C0006"),
    ("BLOCKED", "FFC7CE", "9C0006"),
    ("MANUAL", "FFEB9C", "9C6500"),
    ("UNKNOWN", "FFEB9C", "9C6500"),
    ("WARNING", "FFEB9C", "9C6500"),
    ("PASS", "C6EFCE", "006100"),
)


def _stat_range(sheet: str, col: str, count: int) -> str:
    """跨表统计范围。空表也给安全范围（A2:A2 → 计数 0），绝不产生 A2:A1 反转引用。"""
    return f"'{sheet}'!{col}2:{col}{max(count + 1, 2)}"


def _highlight_status(ws, col: str, data_rows: int) -> None:
    """按状态值给指定列加条件格式（打开工作簿即渲染，不依赖宏）。"""
    rng = f"{col}2:{col}{max(data_rows + 1, 2)}"
    for value, bg, fg in _STATUS_HIGHLIGHTS:
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal", formula=[f'"{value}"'],
            fill=PatternFill("solid", start_color=bg, end_color=bg),
            font=Font(color=fg)))


def _write_cover_stats(ws, rules, queries, evidences, reviews) -> None:
    """封面与汇总追加"状态统计"区块：真实 Excel 公式，打开即重算、随数据行数联动。"""
    ws.append([])
    title = ws.cell(row=ws.max_row + 1, column=1,
                    value="状态统计（Excel 公式实时联动，随明细数据重算）")
    title.font = HEADER_FONT
    header_row = ws.max_row + 1
    for col, text in ((1, "统计项"), (2, "数量"), (3, "口径说明")):
        ws.cell(row=header_row, column=col, value=text).font = HEADER_FONT

    rule_rng = lambda col: _stat_range("条款核查结论", col, len(rules))  # noqa: E731
    query_rng = lambda col: _stat_range("数据源查询日志", col, len(queries))  # noqa: E731

    def add(label, formula, note):
        ws.cell(row=ws.max_row + 1, column=1, value=label)
        ws.cell(row=ws.max_row, column=2, value=formula)      # 公式：程序构造，直写
        ws.cell(row=ws.max_row, column=3, value=_cell(note))  # 文本：一律走消毒

    first_status_row = ws.max_row + 1
    add("条款核查项数", f"=COUNTA({rule_rng('A')})", "含'不适用'条款")
    add("触发否决条款（FAIL）", f'=COUNTIF({rule_rng("C")},"FAIL")',
        "A/B 级官方证据触发否决")
    add("风险提示（WARNING）", f'=COUNTIF({rule_rng("C")},"WARNING")', "发现风险，不足以否决")
    add("通过（PASS）", f'=COUNTIF({rule_rng("C")},"PASS")', "查询成功且未发现触发记录")
    add("需人工（MANUAL）", f'=COUNTIF({rule_rng("C")},"MANUAL")', "验证码/登录/复核——不是正常")
    add("查询失败（ERROR）", f'=COUNTIF({rule_rng("C")},"ERROR")', "绝不是'无异常'")
    add("查询超时（TIMEOUT）", f'=COUNTIF({rule_rng("C")},"TIMEOUT")', "绝不是'无异常'")
    add("访问受限（BLOCKED）", f'=COUNTIF({rule_rng("C")},"BLOCKED")', "绝不是'无异常'")
    add("证据不足（UNKNOWN）", f'=COUNTIF({rule_rng("C")},"UNKNOWN")', "绝不是'正常'")
    add("不适用（NOT_APPLICABLE）", f'=COUNTIF({rule_rng("C")},"NOT_APPLICABLE")',
        "行业/集团不匹配未查询")
    add("异常与待人工合计", f"=SUM(B{first_status_row + 4}:B{first_status_row + 8})",
        "MANUAL/ERROR/TIMEOUT/BLOCKED/UNKNOWN——红线：绝不视作'无异常'")
    add("数据源查询项数", f"=COUNTA({query_rng('A')})", "本批次实际发起的源查询")
    add("数据源异常/待人工",
        f'=COUNTIF({query_rng("B")},"MANUAL")+COUNTIF({query_rng("B")},"ERROR")'
        f'+COUNTIF({query_rng("B")},"TIMEOUT")+COUNTIF({query_rng("B")},"BLOCKED")'
        f'+COUNTIF({query_rng("B")},"UNKNOWN")', "五个状态任一出现即需人工过目")
    add("证据条数", f"=COUNTA({_stat_range('证据清单', 'A', len(evidences))})",
        "带 SHA-256 的原始证据")
    add("人工复核记录数", f"=COUNTA({_stat_range('人工复核记录', 'A', len(reviews))})",
        "复核流写入的审计记录")
    for col, width in ((1, 26), (2, 10), (3, 46)):
        ws.column_dimensions[ws.cell(row=header_row, column=col).column_letter].width = width

_SHEET_TITLES = (
    "封面与汇总", "项目信息", "企业信息", "条款核查结论", "数据源查询日志",
    "发现明细", "证据清单", "人工复核记录", "数据源注册表", "状态口径说明", "免责与合规声明",
)


def _connect(db_path):
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    return conn


def _latest_run(conn, pc_id):
    pc = conn.execute(
        "SELECT * FROM project_companies WHERE id = ?", (pc_id,)).fetchone()
    if pc is None:
        raise ValueError(f"project_companies id={pc_id} 不存在")
    run = conn.execute(
        "SELECT * FROM check_runs WHERE project_id = ? AND company_id = ? "
        "AND finished_at IS NOT NULL ORDER BY id DESC LIMIT 1",
        (pc["project_id"], pc["company_id"])).fetchone()
    project = conn.execute("SELECT * FROM projects WHERE id = ?", (pc["project_id"],)).fetchone()
    company = conn.execute("SELECT * FROM companies WHERE id = ?", (pc["company_id"],)).fetchone()
    return pc, run, project, company


def _write_table(ws, headers, rows):
    ws.append([_cell(h) for h in headers])
    for c in ws[1]:
        c.font = HEADER_FONT
    for r in rows:
        ws.append([_cell(v) for v in r])


def export_excel(db_path: str | Path, pc_id: int, out_path: str | Path) -> Path:
    conn = _connect(db_path)
    try:
        pc, run, project, company = _latest_run(conn, pc_id)
        run_id = run["run_id"] if run else None

        rules = [dict(r) for r in conn.execute(
            "SELECT rule_id, scope, status, reasons_json FROM rule_results "
            "WHERE run_id = ? ORDER BY id", (run_id,))] if run_id else []
        queries = [dict(q) for q in conn.execute(
            "SELECT id, source_id, status, query_url, queried_at, raw_json FROM source_queries "
            "WHERE run_id = ? ORDER BY id", (run_id,))] if run_id else []
        findings = [dict(f) for f in conn.execute(
            "SELECT sq.source_id, f.kind, f.grade, f.description, f.start_date, f.end_date, "
            "f.attrs_json FROM findings f JOIN source_queries sq ON f.query_id = sq.id "
            "WHERE sq.run_id = ? ORDER BY f.id", (run_id,))] if run_id else []
        evidences = [dict(e) for e in conn.execute(
            "SELECT id, source_id, url, kind, sha256, captured_at, key_text FROM evidence "
            "WHERE query_id IN (SELECT id FROM source_queries WHERE run_id = ?) ORDER BY id",
            (run_id,))] if run_id else []
        reviews = [dict(r) for r in conn.execute(
            "SELECT reviewer, decision, note, reviewed_at FROM manual_reviews "
            "WHERE run_id = ? ORDER BY id", (run_id,))] if run_id else []
        # 注册表以包内 YAML 为权威（DB source_registry 表无写入方，P6 复核轮修复）
        import yaml

        reg_data = yaml.safe_load(
            (Path(__file__).resolve().parents[1] / "config" / "sources_registry.yaml")
            .read_text(encoding="utf-8")) or {}
        registry = sorted(reg_data.get("sources") or [], key=lambda x: x.get("id", ""))
    finally:
        conn.close()

    import json as _json
    for r in rules:
        r["reasons"] = _json.loads(r.pop("reasons_json") or "[]")

    wb = Workbook()
    ws = wb.active

    # 1 封面与汇总
    ws.title = _SHEET_TITLES[0]
    rows = [
        ("报告", "投标人资格核查明细表"),
        ("生成版本", f"bqc {__version__}"),
        ("项目", project["name"]),
        ("企业", f'{company["name"]}' + (f'（{company["uscc"]}）' if company["uscc"] else "")),
        ("核查基准日", project["base_date"]),
        ("数据来源", _source_provenance(run)),
        ("总体结论", f'{run["overall_status"]} · {report_label(Status(run["overall_status"]))}'
         if run and run["overall_status"] else "尚无完整核查运行"),
        ("业务判断（decision）", run["decision_status"] if run else None),
        ("数据获取（data）", f'{run["data_status"]} · {report_label(Status(run["data_status"]))}'
         if run and run["data_status"] else None),
        ("需人工复核", "是" if run and run["manual_required"] else "否"),
        ("本次批次 run_id", run["run_id"] if run else None),
        ("批次开始/完成", f'{run["started_at"]} / {run["finished_at"]}' if run else None),
        ("声明", "本表为机器核查明细；'查询失败/超时/待人工'绝不代表'无异常'；"
               "'未查到'≠'确认不存在'。"),
    ]
    _write_table(ws, ["项目", "内容"], rows)
    _write_cover_stats(ws, rules, queries, evidences, reviews)

    # 2 项目信息
    ws = wb.create_sheet(_SHEET_TITLES[1])
    _write_table(ws, ["id", "名称", "省", "市", "行业", "招标人集团", "基准日", "近几年", "启用条款"],
                 [(project["id"], project["name"], project["province"], project["city"],
                   project["industry"], project["owner_group"], project["base_date"],
                   project["years_back"], project["terms"])])

    # 3 企业信息
    ws = wb.create_sheet(_SHEET_TITLES[2])
    _write_table(ws, ["id", "企业名称", "统一社会信用代码", "注册地省", "本次核查状态", "总体结论"],
                 [(company["id"], company["name"], company["uscc"],
                   company["registered_province"], pc["status"], pc["overall_status"])])

    # 4 条款核查结论
    ws = wb.create_sheet(_SHEET_TITLES[3])
    rule_rows = []
    for r in rules:
        label = ("不适用（未启用条款）" if r["status"] == "NOT_APPLICABLE"
                 else report_label(Status(r["status"])))
        rule_rows.append((r["rule_id"], r.get("scope"), r["status"], label,
                          "；".join(r["reasons"])))
    _write_table(ws, ["规则", "层级", "状态", "结论用语", "判定依据"], rule_rows)
    _highlight_status(ws, "C", len(rule_rows))

    # 5 数据源查询日志
    ws = wb.create_sheet(_SHEET_TITLES[4])
    q_rows = []
    for q in queries:
        label = ("不适用（行业/集团不匹配）" if q["status"] == "NOT_APPLICABLE"
                 else report_label(Status(q["status"])))
        q_rows.append((q["source_id"], q["status"], label, q["queried_at"],
                       q["query_url"], _json.loads(q["raw_json"] or "{}").get("note", "")))
    _write_table(ws, ["数据源", "状态", "结论用语", "查询时间", "入口", "备注"], q_rows)
    _highlight_status(ws, "B", len(q_rows))

    # 6 发现明细
    ws = wb.create_sheet(_SHEET_TITLES[5])
    _write_table(ws, ["数据源", "类型", "证据等级", "描述", "起始日", "截止日", "属性"],
                 [(f["source_id"], f["kind"], f["grade"], f["description"],
                   f["start_date"], f["end_date"], f["attrs_json"]) for f in findings])

    # 7 证据清单
    ws = wb.create_sheet(_SHEET_TITLES[6])
    _write_table(ws, ["id", "数据源", "URL", "类型", "SHA-256", "采集时间", "摘要"],
                 [(e["id"], e["source_id"], e["url"], e["kind"], e["sha256"],
                   e["captured_at"], e["key_text"]) for e in evidences])

    # 8 人工复核记录
    ws = wb.create_sheet(_SHEET_TITLES[7])
    _write_table(ws, ["复核人", "结论", "备注", "时间"],
                 [(r["reviewer"], r["decision"], r["note"], r["reviewed_at"]) for r in reviews])

    # 9 数据源注册表
    ws = wb.create_sheet(_SHEET_TITLES[8])
    _write_table(ws, ["id", "名称", "层级", "省", "集团", "官方入口", "查询接口",
                      "自动化模式", "启用"],
                 [(r.get("id", ""), r.get("name", ""), r.get("level", ""),
                   r.get("province") or "", r.get("owner_group") or "",
                   r.get("official_home") or "", r.get("query_url") or "",
                   r.get("automation_mode", ""),
                   "是" if r.get("enabled") else "否") for r in registry])

    # 10 状态口径说明（红线：失败绝不写成无异常）
    ws = wb.create_sheet(_SHEET_TITLES[9])
    legend = [
        ("PASS", report_label(Status.PASS), "查询成功且未发现触发记录"),
        ("WARNING", report_label(Status.WARNING), "发现风险，不足以否决"),
        ("FAIL", report_label(Status.FAIL), "存在 A/B 级官方证据触发否决条款"),
        ("MANUAL", report_label(Status.MANUAL), "需人工验证码/登录/复核——不是正常"),
        ("NO_DATA", report_label(Status.NO_DATA), "查询成功但未检索到记录——不是'确认不存在'"),
        ("ERROR", report_label(Status.ERROR), "查询失败——绝不是'无异常'"),
        ("TIMEOUT", report_label(Status.TIMEOUT), "查询超时——绝不是'无异常'"),
        ("BLOCKED", report_label(Status.BLOCKED), "访问被限制——绝不是'无异常'"),
        ("UNKNOWN", report_label(Status.UNKNOWN), "证据不足无法确认——绝不是'正常'"),
        ("NOT_APPLICABLE", "不适用", "数据源与本项目行业/集团不匹配，未查询（不是查询结果）"),
    ]
    _write_table(ws, ["状态", "报告用语", "含义"], legend)
    ws.append([])
    ws.append(["硬规则", "ERROR / TIMEOUT / BLOCKED / MANUAL / UNKNOWN 永不自动算作 PASS；"
               "本表任何单元格都不把失败状态表述为'正常/无异常'。"])

    # 11 免责与合规声明
    ws = wb.create_sheet(_SHEET_TITLES[10])
    _write_table(ws, ["声明"], [
        ("本工具仅做公开信息的自动查询与整理，不破解验证码、不绕过反自动化、不伪造身份。"),
        ("'未查到'与'确认不存在'严格区分；机器结论不替代人工复核与招标文件条款解释。"),
        ("遇到 MANUAL/ERROR/TIMEOUT/BLOCKED/UNKNOWN 状态的，必须人工核实后方可使用本表结论。"),
        ("本表不含任何企业内部受限名单、用户查询记录与登录凭证；证据原件存本地证据目录。"),
    ])

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
