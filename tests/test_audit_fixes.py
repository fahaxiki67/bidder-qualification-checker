"""审计整改回归测试（2026-09-05 代码审计）。

每一条对应一处已修复的缺陷，命名即结论：这些用例在修复前必然失败。
红线：任何"静默降级/静默成功"的行为都必须在此被钉死。
"""
from __future__ import annotations

import importlib
import json
import os
import sqlite3
import sys
import tempfile
from datetime import date
from pathlib import Path

import pytest
import yaml

from app.core import evidence as ev
from app.core import runner as R
from app.core.db import connect, init_db
from app.core.models import Company, Finding, Project
from app.core.runner import run_check
from app.core.rules import BidRestrictionRule, RuleEngine
from app.sources.mock import SCENARIOS

SUBJ = {"subject_name": "测试建筑有限公司", "subject_uscc": "91510000TEST0000XX"}


# ---------------------------------------------------------------- 公用夹具
@pytest.fixture()
def real_env(tmp_path, monkeypatch):
    """多源真实链路环境：3 个源各有不同响应原文 + 1 个源无原文（尾源）。"""
    reg = tmp_path / "reg.yaml"
    reg.write_text(yaml.safe_dump({"sources": [
        {"id": "creditchina", "name": "信用中国", "level": "national", "automation_mode": "auto",
         "official_home": "https://www.creditchina.gov.cn/",
         "query_url": "https://www.creditchina.gov.cn/q",
         "evidence_grade": "A", "adapter": "app.sources.national.creditchina"},
        {"id": "pcczdc", "name": "破产重整", "level": "national", "automation_mode": "auto",
         "official_home": "https://pccz.court.gov.cn/",
         "query_url": "https://pccz.court.gov.cn/q",
         "evidence_grade": "A", "adapter": "app.sources.national.pcczdc"},
        {"id": "mem_safety_credit", "name": "应急管理部", "level": "national",
         "automation_mode": "auto", "official_home": "https://www.mem.gov.cn/",
         "query_url": "https://www.mem.gov.cn/q",
         "evidence_grade": "A", "adapter": "app.sources.national.mem"},
    ]}, allow_unicode=True), encoding="utf-8")
    app_yaml = tmp_path / "app.yaml"
    app_yaml.write_text("nightly_mock_only: false\n", encoding="utf-8")
    monkeypatch.setattr(R, "REGISTRY_YAML", reg)
    monkeypatch.setattr(R, "APP_YAML", app_yaml)
    db = tmp_path / "data" / "t.sqlite3"
    init_db(db)
    conn = connect(db)
    cur = conn.execute(
        "INSERT INTO projects (name, industry, base_date, years_back, terms) "
        "VALUES ('审计项目','建筑','2026-09-05',3,'条款1,条款2,条款3,条款4')")
    pid = cur.lastrowid
    cur = conn.execute("INSERT INTO companies (name, uscc) VALUES ('测试建筑有限公司','91510000TEST0000XX')")
    cid = cur.lastrowid
    cur = conn.execute("INSERT INTO project_companies (project_id, company_id) VALUES (?,?)", (pid, cid))
    pcid = cur.lastrowid
    conn.commit()
    conn.close()
    return db, pcid


#: 源 id → 其 query_url 中出现的域名片段（用于假传输层按源分流）
URL_HINTS = {"creditchina": "creditchina.gov.cn",
             "pcczdc": "pccz.court.gov.cn",
             "mem_safety_credit": "mem.gov.cn"}


def _mk_get(texts: dict[str, str]):
    def _get(url, timeout):
        for key, body in texts.items():
            if URL_HINTS[key] in url:
                return 200, body
        return 404, ""
    return _get


# ------------------------------------------------- 缺陷 1：证据串号/丢证
def test_evidence_is_per_source_not_last_source(real_env):
    """修复前：所有源共用最后一个 out → 证据串号（且哈希自校验查不出来）。"""
    db, pcid = real_env
    texts = {
        "creditchina": json.dumps({"result": [{**SUBJ, "penalty_content": "限制投标",
                                              "authority_level": "province",
                                              "start_date": "2026-08-05", "end_date": "2027-08-05"}]}),
        "pcczdc": "RAW-PCCZDC-唯一原文",
        "mem_safety_credit": "RAW-MEM-唯一原文",
    }
    run_check(db, pcid, real_sources=True, get=_mk_get(texts))

    conn = sqlite3.connect(str(db))
    conn.row_factory = sqlite3.Row
    rows = conn.execute("SELECT source_id, file_path FROM evidence ORDER BY id").fetchall()
    conn.close()
    bodies = {r["source_id"]: Path(r["file_path"]).read_text(encoding="utf-8") for r in rows}
    assert set(bodies) == set(texts), f"证据未逐源落盘：{set(bodies)}"
    for sid, body in bodies.items():
        assert body == texts[sid], f"{sid} 的证据内容被别的源覆盖（串号）"


def test_evidence_not_lost_when_last_source_has_no_raw_text(real_env):
    """修复前：最后一个源无响应原文 → 前面所有源的证据全被跳过（证据丢失）。"""
    db, pcid = real_env
    texts = {
        "creditchina": "RAW-CREDITCHINA",
        "mem_safety_credit": "RAW-MEM",
        "pcczdc": "",  # 尾源无原文（MANUAL/未复核场景常见）
    }
    run_check(db, pcid, real_sources=True, get=_mk_get(texts))
    conn = sqlite3.connect(str(db))
    got = {r[0] for r in conn.execute("SELECT source_id FROM evidence")}
    conn.close()
    assert got == {"creditchina", "mem_safety_credit"}, f"证据丢失：{got}"


def test_evidence_carries_source_grade(real_env):
    """证据等级随源登记（决定能否支持 FAIL），不得为 NULL。"""
    db, pcid = real_env
    run_check(db, pcid, real_sources=True,
              get=_mk_get({"creditchina": "A级源原文", "pcczdc": "", "mem_safety_credit": ""}))
    conn = sqlite3.connect(str(db))
    grade = conn.execute("SELECT grade FROM evidence WHERE source_id='creditchina'").fetchone()[0]
    conn.close()
    assert grade == "A"


# --------------------------------------- 缺陷 2：大证据截断后哈希必然不符
def test_truncated_evidence_verifies_clean(tmp_path, monkeypatch):
    """修复前：先算哈希再追加截断说明 → 未篡改的证据被自己的校验判为损坏。"""
    db = tmp_path / "d.sqlite3"
    init_db(db)
    monkeypatch.setattr(ev, "MAX_EVIDENCE_BYTES", 1000)
    eid, fpath, digest = ev.save_evidence(db, source_id="x", url="https://e.com",
                                          raw_text="某" * 5000, kind="raw_response")
    body = Path(fpath).read_text(encoding="utf-8")
    assert "已截断" in body
    assert ev.sha256_text(body) == digest
    ok, broken = ev.verify_evidence(db, eid)
    assert broken == [], f"未篡改的证据被判损坏：{broken}"
    assert ok == 1


# ------------------------------------------------ 缺陷 3：Excel 公式注入
def test_excel_neutralises_formula_text(tmp_path):
    """报告内容含第三方文本，= 开头的字符串不得原样落单元格。"""
    from app.reports.excel import export_excel

    db = tmp_path / "d.sqlite3"
    init_db(db)
    conn = connect(db)
    cur = conn.execute("INSERT INTO projects (name, base_date, years_back) VALUES ('P','2026-09-05',3)")
    pid = cur.lastrowid
    cur = conn.execute("INSERT INTO companies (name) VALUES ('=1+1')")
    cid = cur.lastrowid
    cur = conn.execute("INSERT INTO project_companies (project_id, company_id) VALUES (?,?)", (pid, cid))
    pcid = cur.lastrowid
    conn.commit()
    conn.close()
    run_check(db, pcid, scenario="clean")

    out = tmp_path / "r.xlsx"
    export_excel(db, pcid, out)
    import openpyxl
    wb = openpyxl.load_workbook(out)
    risky = [c.value for ws in wb for row in ws.iter_rows() for c in row
             if isinstance(c.value, str) and c.value[:1] in ("=", "+", "-", "@")]
    assert risky == [], f"存在可直接被当公式执行的单元格：{risky}"
    texts = [c.value for ws in wb for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert any(t.startswith("'=1+1") for t in texts), "危险值应保留内容并强制按文本处理"


# ------------------------------------- 缺陷 4：PDF 长文本溢出 / 特殊字符
def test_pdf_wraps_long_text_and_escapes_xml(tmp_path):
    """修复前：Table 单元格为纯字符串 → 长中文不折行溢出页面；含 & < 时解析失败。"""
    import pypdf
    from app.reports.pdf import export_pdf

    long_name = "四川某超长企业名称集团股份有限公司" * 6
    db = tmp_path / "d.sqlite3"
    init_db(db)
    conn = connect(db)
    cur = conn.execute("INSERT INTO projects (name, base_date, years_back) VALUES (?, '2026-09-05', 3)",
                       (long_name,))
    pid = cur.lastrowid
    cur = conn.execute("INSERT INTO companies (name) VALUES (?)", (long_name + " & <测试> ",))
    cid = cur.lastrowid
    cur = conn.execute("INSERT INTO project_companies (project_id, company_id) VALUES (?,?)", (pid, cid))
    pcid = cur.lastrowid
    conn.commit()
    conn.close()
    run_check(db, pcid, scenario="bid_ban")

    out = tmp_path / "r.pdf"
    export_pdf(db, pcid, out)  # 修复前此处抛 XML 解析错误
    raw = "\n".join(p.extract_text() for p in pypdf.PdfReader(str(out)).pages)
    # 折行后文本被 Paragraph 切成多行 → 去换行再比对；能整段保留即证明没有溢出裁切
    text = raw.replace("\n", "")
    assert long_name in text
    assert raw.count("\n") > 10, "长文本应已被折行为多行（修复前是单行溢出页面）"
    assert "判定依据" in text          # PDF 必须带判定依据（此前只有状态、没有理由）
    assert "触发否决条款" in text
    assert "证据清单" in text


# -------------------------- 缺陷 5：mock 链路把"没查"的源记成 PASS
def test_mock_run_does_not_claim_unqueried_sources_succeeded(tmp_path):
    """修复前：10 个源全部记 PASS（查询成功）——报告凭空声称查过所有官方源。"""
    db = tmp_path / "d.sqlite3"
    init_db(db)
    conn = connect(db)
    cur = conn.execute("INSERT INTO projects (name, province, industry, owner_group, base_date,"
                       " years_back, terms) VALUES ('演示','广东','建筑','powerchina','2026-09-05',3,NULL)")
    pid = cur.lastrowid
    cur = conn.execute("INSERT INTO companies (name, registered_province) VALUES ('某公司','四川')")
    cid = cur.lastrowid
    cur = conn.execute("INSERT INTO project_companies (project_id, company_id) VALUES (?,?)", (pid, cid))
    pcid = cur.lastrowid
    conn.commit()
    conn.close()

    assert run_check(db, pcid, scenario="bid_ban") == "FAIL"
    conn = sqlite3.connect(str(db))
    conn.row_factory = sqlite3.Row
    rows = [dict(r) for r in conn.execute(
        "SELECT source_id, status, raw_json FROM source_queries ORDER BY id")]
    conn.close()
    assert len(rows) > 1
    passed = [r for r in rows if r["status"] == "PASS"]
    assert len(passed) <= 1, f"演示链路只有一个主源返回数据，其余不应记 PASS：{[r['source_id'] for r in passed]}"
    for r in rows:
        if r["status"] == "NO_DATA":
            assert "未实际查询" in json.loads(r["raw_json"] or "{}").get("note", "")


# ------------------------------ 缺陷 6：未知演示场景静默按"无发现"处理
def test_unknown_scenario_raises(tmp_path):
    db = tmp_path / "d.sqlite3"
    init_db(db)
    conn = connect(db)
    cur = conn.execute("INSERT INTO projects (name, base_date, years_back) VALUES ('P','2026-09-05',3)")
    pid = cur.lastrowid
    cur = conn.execute("INSERT INTO companies (name) VALUES ('C')")
    cid = cur.lastrowid
    cur = conn.execute("INSERT INTO project_companies (project_id, company_id) VALUES (?,?)", (pid, cid))
    pcid = cur.lastrowid
    conn.commit()
    conn.close()
    with pytest.raises(ValueError, match="未知演示场景"):
        run_check(db, pcid, scenario="bidd_ban")   # 拼错
    for ok in SCENARIOS:
        run_check(db, pcid, scenario=ok)           # 全部合法场景仍可用


def test_web_rejects_unknown_scenario(tmp_path, monkeypatch):
    monkeypatch.setenv("BQC_DB", str(tmp_path / "b.sqlite3"))
    import app.web.server as server
    importlib.reload(server)
    from fastapi.testclient import TestClient
    c = TestClient(server.app)
    r = c.post("/projects", data={"project_name": "P", "base_date": "2026-09-05",
                                  "company_name": "C", "scenario": "bidd_ban"},
               follow_redirects=False)
    assert r.status_code == 400


# --------------------------- 缺陷 7：years_back 采集了却完全不参与判断
def test_years_back_window_is_surfaced_in_reasons():
    """窗口外的记录必须在判定依据里显式提示（不自动降级 FAIL，只要求人工复核口径）。"""
    project = Project(name="P", base_date=date(2026, 9, 4), years_back=3)
    old = Finding(kind="penalty_bid_restriction", source_id="creditchina", grade="A",
                  description="8 年前的限制投标处罚（未载明解除日期）",
                  start_date=date(2018, 1, 1), end_date=None,
                  attrs={"authority_level": "province"})
    company = Company(name="测试建筑有限公司", uscc="91510000TEST0000XX")

    def _bid_reasons(findings):
        res = RuleEngine().run_all(findings, project, company)
        return [r for r in res if r.rule_id == "rule_bid_restriction"][0]

    out = _bid_reasons([old])
    assert out.status == "FAIL", "窗口外记录不得被自动抹掉（宁可多提示，不可漏否决）"
    assert any("窗口" in r for r in out.reasons)

    fresh = Finding(kind="penalty_bid_restriction", source_id="creditchina", grade="A",
                    description="近期限投标", start_date=date(2026, 1, 1), end_date=None,
                    attrs={"authority_level": "province"})
    assert not any("窗口" in r for r in _bid_reasons([fresh]).reasons)

    # 无日期记录不得被当成超窗口（宁可提示，不可据此放宽）
    undated = Finding(kind="penalty_bid_restriction", source_id="creditchina", grade="A",
                      description="未载明日期", attrs={"authority_level": "province"})
    assert not any("窗口" in r for r in _bid_reasons([undated]).reasons)


def test_window_note_reaches_rule_engine_results():
    project = Project(name="P", base_date=date(2026, 9, 4), years_back=3, terms=("条款1",))
    old = Finding(kind="penalty_bid_restriction", source_id="creditchina", grade="A",
                  description="旧处罚", start_date=date(2015, 1, 1), end_date=date(2016, 1, 1),
                  attrs={"authority_level": "province"})
    results = RuleEngine().run_all([old], project, Company(name="C"))
    bid = [r for r in results if r.rule_id == "rule_bid_restriction"][0]
    assert any("窗口" in r for r in bid.reasons)


# --------------------------------------- 缺陷 8：本地 UI 跨站表单写保护
def test_cross_site_post_rejected(tmp_path, monkeypatch):
    monkeypatch.setenv("BQC_DB", str(tmp_path / "b.sqlite3"))
    import app.web.server as server
    importlib.reload(server)
    from fastapi.testclient import TestClient
    c = TestClient(server.app)
    data = {"project_name": "伪造", "base_date": "2026-09-05", "company_name": "C", "scenario": "clean"}
    evil = {"Origin": "http://evil.example", "Referer": "http://evil.example/form"}
    assert c.post("/projects", data=data, headers=evil, follow_redirects=False).status_code == 403
    same = {"Origin": "http://testserver", "Referer": "http://testserver/"}
    assert c.post("/projects", data=data, headers=same, follow_redirects=False).status_code == 303
    assert c.post("/projects", data=data, follow_redirects=False).status_code in (303, 400)


# --------------------------- 缺陷 9：serve 绑定非回环地址（无鉴权暴露）
def test_serve_refuses_non_loopback_without_flag(monkeypatch, capsys):
    from app import main as cli

    started = {}
    monkeypatch.setattr("uvicorn.run", lambda app, host, port: started.setdefault("host", host))
    assert cli.main(["serve", "--host", "0.0.0.0"]) == 2
    assert "host" not in started
    assert cli.main(["serve", "--host", "0.0.0.0", "--allow-lan"]) == 0
    assert started["host"] == "0.0.0.0"
    assert "warn" in capsys.readouterr().err


# ---------------------- 缺陷 10：import app.web.server 的建库副作用
def test_importing_web_server_has_no_side_effect(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    monkeypatch.setenv("BQC_DB", str(tmp_path / "b.sqlite3"))
    sys.modules.pop("app.web.server", None)
    import app.web.server  # noqa: F401
    assert not (tmp_path / "data").exists(), "导入模块不得在当前工作目录建库"
    assert not (tmp_path / "b.sqlite3").exists()


# --------------------------------------- 缺陷 11：SQLite 并发（WAL/等待）
def test_connect_uses_wal_and_busy_timeout(tmp_path):
    db = tmp_path / "d.sqlite3"
    conn = connect(db)
    try:
        mode = conn.execute("PRAGMA journal_mode").fetchone()[0]
        assert mode.lower() == "wal"
        assert conn.execute("PRAGMA busy_timeout").fetchone()[0] >= 5000
        assert conn.execute("PRAGMA foreign_keys").fetchone()[0] == 1
    finally:
        conn.close()


# ------------------------------------------ CLI 端到端（main.py 此前无测试）
def test_cli_entrypoints(tmp_path, capsys):
    from app import main as cli

    with pytest.raises(SystemExit):  # argparse --version 直接退出
        cli.main(["--version"])
    db = tmp_path / "data" / "b.sqlite3"
    assert cli.main(["init-db", "--db", str(db)]) == 0
    assert db.is_file()
    assert cli.main(["verify-evidence", "--db", str(db)]) == 0
    assert cli.main(["report", "999", "--db", str(db), "--excel", str(tmp_path / "x.xlsx")]) == 2
    missing = tmp_path / "nope.json"
    assert cli.main(["import-bans", str(missing), "--db", str(db)]) == 2
    capsys.readouterr()


# -------------------------- 整改加强：数据来源口径（mock ≠ 真实官方查询）
def _mk_pc(db, terms="条款1,条款2,条款3,条款4"):
    conn = connect(db)
    cur = conn.execute("INSERT INTO projects (name, base_date, years_back, terms) "
                       "VALUES ('P','2026-09-05',3,?)", (terms,))
    pid = cur.lastrowid
    cur = conn.execute("INSERT INTO companies (name) VALUES ('演示企业')")
    cid = cur.lastrowid
    cur = conn.execute("INSERT INTO project_companies (project_id, company_id) VALUES (?,?)", (pid, cid))
    pcid = cur.lastrowid
    conn.commit()
    conn.close()
    return pcid


def test_reports_mark_mock_runs_as_demo_data(tmp_path):
    """mock 演示链路必须在报告显著位置标注，不能被当成真实官方查询结论。"""
    import openpyxl
    import pypdf
    from app.reports.excel import export_excel
    from app.reports.pdf import export_pdf

    db = tmp_path / "d.sqlite3"
    init_db(db)
    pcid = _mk_pc(db)
    run_check(db, pcid, scenario="clean")

    xlsx = tmp_path / "r.xlsx"
    export_excel(db, pcid, xlsx)
    cover = [c.value for row in openpyxl.load_workbook(xlsx)["封面与汇总"].iter_rows() for c in row]
    assert any(isinstance(v, str) and "mock 演示链路" in v for v in cover)

    pdf = tmp_path / "r.pdf"
    export_pdf(db, pcid, pdf)
    text = "".join(p.extract_text() for p in pypdf.PdfReader(str(pdf)).pages)
    assert "mock 演示链路" in text


def test_result_page_flags_mock_run(tmp_path, monkeypatch):
    monkeypatch.setenv("BQC_DB", str(tmp_path / "b.sqlite3"))
    import app.web.server as server
    importlib.reload(server)
    from fastapi.testclient import TestClient
    c = TestClient(server.app)
    r = c.post("/projects", data={"project_name": "演示", "base_date": "2026-09-05",
                                  "company_name": "某公司", "scenario": "clean"},
               follow_redirects=False)
    assert r.status_code == 303
    page = c.get(r.headers["location"]).text
    assert "mock 演示链路" in page
    assert "不得作为核查结论使用" in page
