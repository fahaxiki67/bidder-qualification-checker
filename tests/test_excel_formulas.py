"""报告内置 Excel 公式案例：封面统计区跨表 COUNTIF/SUM 联动 + 状态列条件格式。

契约（与 _cell() 消毒互补，两条路径互斥）：
- 公式只允许出现在"封面与汇总"统计区块，全部由程序按行号构造、零第三方文本；
- 公式必须是真公式单元格（openpyxl data_type=='f'），Excel/WPS/LibreOffice 打开即重算；
- 数据路径依旧零公式：第三方危险前缀文本一律消毒为文本（见 test_audit_fixes）。
"""
import json

import pytest
import yaml

from app.core import runner
from app.core.db import connect, init_db
from app.core.runner import run_check
from app.reports.excel import export_excel

SUBJ = {"subject_name": "测试建筑有限公司", "subject_uscc": "91510000TEST0000XX"}

REGISTRY = {"sources": [
    {"id": "creditchina", "name": "信用中国", "level": "national",
     "automation_mode": "auto",
     "official_home": "https://www.creditchina.gov.cn/",
     "query_url": "https://www.creditchina.gov.cn/q",
     "adapter": "app.sources.national.creditchina"}]}

FAIL_FIXTURE = {"result": [{**SUBJ, "penalty_content": "省级住建主管部门限制投标一年",
                            "authority_level": "province",
                            "start_date": "2026-08-05", "end_date": "2027-08-05"}]}


@pytest.fixture()
def env(tmp_path, monkeypatch):
    reg = tmp_path / "sources_registry.yaml"
    reg.write_text(yaml.safe_dump(REGISTRY, allow_unicode=True), encoding="utf-8")
    app_yaml = tmp_path / "app.yaml"
    app_yaml.write_text("nightly_mock_only: false\n", encoding="utf-8")
    monkeypatch.setattr(runner, "REGISTRY_YAML", reg)
    monkeypatch.setattr(runner, "APP_YAML", app_yaml)
    db = tmp_path / "data" / "t.sqlite3"
    init_db(db)

    def mk_pc(name, terms="条款1,条款2,条款3,条款4"):
        conn = connect(db)
        cur = conn.execute("INSERT INTO projects (name, industry, base_date, years_back, terms) "
                           "VALUES (?,?,?,?,?)", (name, "建筑", "2026-09-05", 3, terms))
        pid = cur.lastrowid
        row = conn.execute("SELECT id FROM companies WHERE uscc = ?",
                           ("91510000TEST0000XX",)).fetchone()
        cid = row[0] if row else conn.execute(
            "INSERT INTO companies (name, uscc) VALUES (?,?)",
            ("测试建筑有限公司", "91510000TEST0000XX")).lastrowid
        cur = conn.execute("INSERT INTO project_companies (project_id, company_id, status) "
                           "VALUES (?,?, 'running')", (pid, cid))
        pcid = cur.lastrowid
        conn.commit()
        conn.close()
        return pcid

    return db, mk_pc, tmp_path


def _formula_cells(wb):
    """全簿所有真公式单元格 [(sheet, cell, value)]。"""
    return [(ws.title, c, c.value) for ws in wb for row in ws.iter_rows() for c in row
            if c.data_type == "f"]


def _detail_rows(ws, col="A"):
    """明细 sheet 的数据行数（去掉表头）。"""
    return sum(1 for row in ws.iter_rows(min_row=2, min_col=1, max_col=1)
               if row[0].value not in (None, ""))


def test_cover_stats_are_real_formulas_whitelisted(env):
    """统计区块是真公式（data_type=='f'）且只允许出现在封面与汇总；范围锚定真实行数。"""
    db, mk_pc, tmp_path = env
    pc = mk_pc("FAIL项目")
    assert run_check(db, pc, real_sources=True,
                     get=lambda url, t: (200, json.dumps(FAIL_FIXTURE))) == "FAIL"
    out = tmp_path / "r.xlsx"
    export_excel(db, pc, out)

    import openpyxl
    wb = openpyxl.load_workbook(out)
    formulas = _formula_cells(wb)
    assert formulas, "封面统计区应存在公式单元格"
    sheets = {s for s, _, _ in formulas}
    assert sheets == {"封面与汇总"}, f"公式只允许出现在封面与汇总，越界：{sheets}"

    cover = wb["封面与汇总"]
    rule_end = 1 + _detail_rows(wb["条款核查结论"])
    fail_f = [v for _, _, v in formulas if '"FAIL"' in v]
    assert f'=COUNTIF(\'条款核查结论\'!C2:C{rule_end},"FAIL")' in fail_f, \
        f"FAIL 计数公式范围应锚定明细行数 {rule_end}，实际：{fail_f}"

    # 合计行：SUM 引用同表五个状态行（MANUAL/ERROR/TIMEOUT/BLOCKED/UNKNOWN）
    sum_f = [v for _, _, v in formulas if v.startswith("=SUM(B")]
    assert len(sum_f) == 1, f"应有且仅有一条 SUM 合计公式：{sum_f}"
    m, n = sum_f[0][5:-1].split(":")   # "=SUM(" 剥头（B 在索引 5）、")" 剥尾 → B21:B25
    lo, hi = int(m[1:]), int(n[1:])
    assert hi - lo == 4, f"合计应恰好覆盖 5 个异常/待人工状态行：{sum_f[0]}"
    # SUM 引用的行本身必须是那五条 COUNTIF
    summed = [cover.cell(row=r, column=2).value for r in range(lo, hi + 1)]
    assert all(isinstance(v, str) and v.startswith('=COUNTIF(') for v in summed), \
        f"SUM 范围内不是状态计数公式：{summed}"


def test_cover_stats_enumerate_all_decision_statuses(env):
    """逐状态计数必须穷举 Status 全集——缺 NO_DATA 时"条款核查项数"与 Σ分项勾稽不齐。

    rules.py 多条路径产出 NO_DATA 规则结果（未检索到记录/主管部门显示正常/他集团禁入），
    封面若不数它，clean 批次会出现 项数 > 九个分项之和 的缺口，且"未检索到"恰是
    红线最强调的状态（≠"确认不存在"），不得从统计区消失。
    """
    db, mk_pc, tmp_path = env
    pc = mk_pc("未检索到项目")
    assert run_check(db, pc, scenario="clean") in ("NO_DATA", "PASS")
    out = tmp_path / "nodata.xlsx"
    export_excel(db, pc, out)

    import openpyxl
    from app.core.status import Status
    wb = openpyxl.load_workbook(out)
    formulas = [v for _, _, v in _formula_cells(wb)]
    rule_end = 1 + _detail_rows(wb["条款核查结论"])
    for st in Status:  # 九态全集，一个都不能少（NOT_APPLICABLE 另有专行）
        expected = f"=COUNTIF('条款核查结论'!C2:C{rule_end},\"{st.value}\")"
        assert expected in formulas, f"封面统计缺 {st.value} 逐状态计数：{expected}"


def test_stats_range_tracks_data_rows(env):
    """明细数据行数变化 → 公式统计范围随之联动（同一 run 前后对比）。"""
    db, mk_pc, tmp_path = env
    pc = mk_pc("联动项目")
    assert run_check(db, pc, scenario="clean") in ("NO_DATA", "PASS")
    out1 = tmp_path / "before.xlsx"
    export_excel(db, pc, out1)

    import openpyxl
    f1 = [v for _, _, v in _formula_cells(openpyxl.load_workbook(out1))
          if '"NOT_APPLICABLE"' in v][0]

    # 向同一 run 再插 3 条规则结果 → 明细变长，重新导出的公式范围必须跟着变
    conn = connect(db)
    run = conn.execute(
        "SELECT run_id, project_id, company_id FROM check_runs "
        "WHERE finished_at IS NOT NULL ORDER BY id DESC LIMIT 1").fetchone()
    for i in range(3):
        conn.execute(
            "INSERT INTO rule_results (project_id, company_id, rule_id, status, "
            "reasons_json, run_id, scope) VALUES (?,?,?,?,?,?,?)",
            (run[1], run[2], f"rule_extra_{i}", "PASS",
             "[]", run[0], "term"))
    conn.commit()
    conn.close()

    out2 = tmp_path / "after.xlsx"
    export_excel(db, pc, out2)
    wb2 = openpyxl.load_workbook(out2)
    f2 = [v for _, _, v in _formula_cells(wb2) if '"NOT_APPLICABLE"' in v][0]
    assert f1 != f2, f"数据变长后公式范围应联动：{f1} vs {f2}"
    end = int(f2.split("!")[1].split(",")[0].split(":")[1][1:])
    assert end == 1 + _detail_rows(wb2["条款核查结论"]), \
        f"公式终点应等于明细行数+1：{f2}"


def test_formula_and_malicious_text_coexist(env):
    """程序公式放行、第三方危险文本消毒——同一文件内互不干扰。"""
    db, mk_pc, tmp_path = env
    pc = mk_pc("恶意名项目")
    conn = connect(db)
    conn.execute("UPDATE companies SET name = '=1+1+cmd' WHERE uscc = ?",
                 ("91510000TEST0000XX",))
    conn.commit()
    conn.close()
    assert run_check(db, pc, scenario="clean") in ("NO_DATA", "PASS")
    out = tmp_path / "r.xlsx"
    export_excel(db, pc, out)

    import openpyxl
    wb = openpyxl.load_workbook(out)
    formulas = _formula_cells(wb)
    assert formulas and all(s == "封面与汇总" for s, _, _ in formulas)
    texts = [c.value for ws in wb for row in ws.iter_rows() for c in row
             if isinstance(c.value, str) and c.data_type == "s"]
    assert any(t.startswith("'=1+1+cmd") for t in texts), "恶意企业名应被消毒为文本"
    clean = [t for t in texts if t.startswith("'=1+1+cmd")]
    assert all(not t.startswith("=1+1+cmd") for t in clean), "消毒后不得以 = 开头"


def test_empty_batch_formulas_safe(env):
    """尚无完整核查运行（run=None）：公式仍在且范围不反转（A2:A1 类非法引用零容忍）。"""
    db, mk_pc, tmp_path = env
    pc = mk_pc("空批次项目")
    out = tmp_path / "empty.xlsx"
    export_excel(db, pc, out)  # 不得抛异常

    import openpyxl
    import re
    wb = openpyxl.load_workbook(out)
    formulas = _formula_cells(wb)
    assert formulas, "空批次也应生成统计公式骨架"
    for _, _, v in formulas:
        for m in re.finditer(r"[A-Z](\d+)", v):
            assert int(m.group(1)) >= 2, f"公式含非法行号（<2，空范围反转）：{v}"
    assert any("COUNTA" in v for _, _, v in formulas)


def test_conditional_formatting_on_status_columns(env):
    """条款核查结论 C 列 / 数据源查询日志 B 列有按状态值的条件格式规则。"""
    db, mk_pc, tmp_path = env
    pc = mk_pc("条件格式项目")
    assert run_check(db, pc, real_sources=True,
                     get=lambda url, t: (200, json.dumps(FAIL_FIXTURE))) == "FAIL"
    out = tmp_path / "cf.xlsx"
    export_excel(db, pc, out)

    import openpyxl
    wb = openpyxl.load_workbook(out)
    checked = 0
    for sheet, col in (("条款核查结论", "C"), ("数据源查询日志", "B")):
        rules = [r for cf in wb[sheet].conditional_formatting
                 for r in cf.rules
                 if str(cf.sqref).startswith(col)]
        values = {r.formula[0].strip('"') for r in rules
                  if r.type == "cellIs" and r.operator == "equal" and r.formula}
        assert {"FAIL", "MANUAL", "PASS"} <= values, f"{sheet}！{col} 列缺状态高亮规则：{values}"
        checked += 1
    assert checked == 2
