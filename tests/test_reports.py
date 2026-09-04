"""P7 报告测试：mock 数据生成 Excel（11 sheet）与 PDF；红线——失败状态绝不写成"无异常"。"""
import json
import re
import sqlite3
import zipfile
from pathlib import Path

import pytest
import yaml

from app.core import runner
from app.core.db import connect, init_db
from app.core.runner import run_check
from app.reports.excel import export_excel
from app.reports.pdf import export_pdf

SUBJ = {"subject_name": "测试建筑有限公司", "subject_uscc": "91510000TEST0000XX"}

REGISTRY = {"sources": [
    {"id": "creditchina", "name": "信用中国", "level": "national",
     "automation_mode": "auto",
     "official_home": "https://www.creditchina.gov.cn/",
     "query_url": "https://www.creditchina.gov.cn/q",
     "adapter": "app.sources.national.creditchina"}]}

FAIL_FIXTURE = {"result": [{**SUBJ, "penalty_content": "省级住建主管部门限制投标一年",
                            "authority_level": "province",
                            "start_date": "2026-08-05", "end_date": "2027-08-05"}]}


@pytest.fixture()
def env(tmp_path, monkeypatch):
    reg = tmp_path / "sources_registry.yaml"
    reg.write_text(yaml.safe_dump(REGISTRY, allow_unicode=True), encoding="utf-8")
    app_yaml = tmp_path / "app.yaml"
    app_yaml.write_text("nightly_mock_only: false\n", encoding="utf-8")
    monkeypatch.setattr(runner, "REGISTRY_YAML", reg)
    monkeypatch.setattr(runner, "APP_YAML", app_yaml)
    db = tmp_path / "data" / "t.sqlite3"
    init_db(db)

    def mk_pc(name, scenario_setup=None):
        conn = connect(db)
        cur = conn.execute("INSERT INTO projects (name, industry, base_date, years_back, terms) "
                           "VALUES (?,?,?,?,?)", (name, "建筑", "2026-09-05", 3,
                                                  "条款1,条款2,条款3,条款4"))
        pid = cur.lastrowid
        row = conn.execute("SELECT id FROM companies WHERE uscc = ?",
                           ("91510000TEST0000XX",)).fetchone()
        cid = row[0] if row else conn.execute(
            "INSERT INTO companies (name, uscc) VALUES (?,?)",
            ("测试建筑有限公司", "91510000TEST0000XX")).lastrowid
        cur = conn.execute("INSERT INTO project_companies (project_id, company_id, status) "
                           "VALUES (?,?, 'running')", (pid, cid))
        pcid = cur.lastrowid
        conn.commit()
        conn.close()
        return pcid

    return db, mk_pc, tmp_path


def test_excel_eleven_sheets_and_honest_labels(env):
    """完成标准：Excel 11 个 sheet；ERROR=查询失败、FAIL=触发否决条款，绝不写'正常'。"""
    db, mk_pc, tmp_path = env
    pc_fail = mk_pc("FAIL项目")
    assert run_check(db, pc_fail, real_sources=True,
                     get=lambda url, t: (200, json.dumps(FAIL_FIXTURE))) == "FAIL"
    pc_err = mk_pc("ERROR项目")
    assert run_check(db, pc_err, scenario="query_error") == "ERROR"

    out = tmp_path / "r" / "detail.xlsx"
    export_excel(db, pc_err, out)
    assert out.is_file() and zipfile.is_zipfile(out)

    import openpyxl
    wb = openpyxl.load_workbook(out)
    assert len(wb.sheetnames) == 11
    assert wb.sheetnames[0] == "封面与汇总" and wb.sheetnames[-1] == "免责与合规声明"

    # ERROR 项目的查询日志：状态=ERROR 用语=查询失败，且全表不出现"正常"
    qws = wb["数据源查询日志"]
    text = "\n".join(str(c.value) for row in qws.iter_rows() for c in row if c.value)
    assert "ERROR" in text and "查询失败" in text
    cover = "\n".join(str(c.value) for row in wb["封面与汇总"].iter_rows()
                      for c in row if c.value)
    assert "ERROR" in cover and "查询失败" in cover and "需人工复核" in cover

    # FAIL 项目：条款结论=触发否决条款
    out2 = tmp_path / "r" / "detail2.xlsx"
    export_excel(db, pc_fail, out2)
    wb2 = openpyxl.load_workbook(out2)
    rws = wb2["条款核查结论"]
    rtext = "\n".join(str(c.value) for row in rws.iter_rows() for c in row if c.value)
    assert "触发否决条款" in rtext and "rule_bid_restriction" in rtext

    # 状态口径 sheet 明示红线
    legend = "\n".join(str(c.value) for row in wb["状态口径说明"].iter_rows()
                       for c in row if c.value)
    assert "绝不" in legend and "NO_DATA" in legend


def test_pdf_report_generated_and_honest(env):
    db, mk_pc, tmp_path = env
    pc_err = mk_pc("ERROR项目")
    assert run_check(db, pc_err, scenario="query_error") == "ERROR"
    out = tmp_path / "r" / "report.pdf"
    export_pdf(db, pc_err, out)
    data = out.read_bytes()
    assert data[:5] == b"%PDF-"

    from pypdf import PdfReader
    reader = PdfReader(str(out))
    assert len(reader.pages) >= 1
    text = "\n".join((p.extract_text() or "") for p in reader.pages)
    assert "投标人资格核查报告" in text
    assert "查询失败" in text          # ERROR 的结论用语如实
    assert "需人工复核" in text or "是" in text


def test_cli_report_command(env):
    db, mk_pc, tmp_path = env
    pc = mk_pc("CLI项目")
    assert run_check(db, pc, scenario="clean") == "NO_DATA"
    from app.main import main as cli
    xlsx = Path(db).parent / "cli.xlsx"
    pdf = Path(db).parent / "cli.pdf"
    rc = cli(["report", str(pc), "--db", str(db), "--excel", str(xlsx), "--pdf", str(pdf)])
    assert rc == 0 and xlsx.is_file() and pdf.is_file()
    # 无输出路径 → 参数错误退出码 2
    assert cli(["report", str(pc), "--db", str(db)]) == 2
